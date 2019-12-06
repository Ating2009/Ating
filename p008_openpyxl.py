import openpyxl

path='C:/Users/Administrator/Desktop/pycharm/lianxi/111.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb['Sheet1']
tuple(ws.rows)
max_row=ws.max_row
n=0
li=[]
for cell in list(ws.columns)[1]:
    n+=1
    if(isinstance(cell.value,int)):
        li.append(n)
for i in range(len(li)-1):
    j=li[i+1]-li[i]-4
    x=0
    for x in range(0,j+1):
        ws.cell(li[i]+x+2,1).value = ws.cell(li[i],2).value
for y in range(0,max_row-li[len(li)-1]-2):
    ws.cell(li[len(li)-1]+2+y,1).value =  ws.cell(li[len(li)-1],2).value
wb.save(path)